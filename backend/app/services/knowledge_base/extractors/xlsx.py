import io

from openpyxl import load_workbook

from app.core.exceptions import ExtractionError
from app.services.knowledge_base.extractors.base import ExtractedSegment
from app.services.llm.base import LLMProvider


async def extract_xlsx(
    content: bytes, filename: str, llm_provider: LLMProvider | None = None
) -> list[ExtractedSegment]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ExtractionError(filename, str(exc)) from exc

    segments: list[ExtractedSegment] = []
    for sheet_number, sheet in enumerate(workbook.worksheets, start=1):
        lines: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                lines.append(" | ".join(cells))
        text = "\n".join(lines).strip()
        if text:
            segments.append(ExtractedSegment(text=text, page=sheet_number, section_title=sheet.title))
    return segments
